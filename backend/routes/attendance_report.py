from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import date, datetime
import io
import os
import sys

# Resolve paths
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.models import db, User, Attendance
from backend.utils.security import role_required

attendance_report_bp = Blueprint('attendance_report', __name__)


def _gather_attendance_data(role_filter, report_date=None, department_id=None):
    """
    Gathers attendance data for a specific role on a specific date.
    Returns a list of dicts and summary counts.
    """
    if report_date is None:
        report_date = date.today()

    if department_id:
        users = User.query.filter_by(role=role_filter, department_id=department_id, is_active=True).all()
    else:
        users = User.query.filter_by(role=role_filter, is_active=True).all()

    rows = []
    present = 0
    absent = 0
    late = 0
    on_leave = 0
    on_duty = 0
    not_marked = 0

    for u in users:
        record = Attendance.query.filter_by(user_id=u.id, date=report_date).first()
        status = record.status if record else "Not Marked"
        check_in = record.check_in_time.strftime('%I:%M %p') if (record and record.check_in_time) else "-"
        check_out = record.check_out_time.strftime('%I:%M %p') if (record and record.check_out_time) else "-"
        hours = record.total_hours if record else 0.0

        if status == 'Present':
            present += 1
        elif status == 'Absent':
            absent += 1
        elif status == 'Late':
            late += 1
        elif status == 'On Leave':
            on_leave += 1
        elif status == 'On Duty':
            on_duty += 1
        else:
            not_marked += 1

        rows.append({
            'id': u.id,
            'username': u.username,
            'department': u.department.name if u.department else 'General',
            'status': status,
            'check_in': check_in,
            'check_out': check_out,
            'total_hours': hours
        })

    summary = {
        'total': len(users),
        'present': present,
        'absent': absent,
        'late': late,
        'on_leave': on_leave,
        'on_duty': on_duty,
        'not_marked': not_marked,
    }

    return rows, summary


@attendance_report_bp.route('/summary', methods=['GET'])
@jwt_required()
@role_required(['Admin', 'Staff'])
def get_attendance_summary():
    """
    Returns JSON summary of today's attendance for both students and staff.
    """
    report_date_str = request.args.get('date')
    if report_date_str:
        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    else:
        report_date = date.today()

    current_user_id = get_jwt_identity()
    current_user = db.session.get(User, int(current_user_id))
    
    department_id = None
    if current_user.role == 'Staff':
        department_id = current_user.department_id

    student_rows, student_summary = _gather_attendance_data('Student', report_date, department_id)
    
    if current_user.role == 'Staff':
        staff_rows = []
        staff_summary = {
            'total': 0, 'present': 0, 'absent': 0, 'late': 0, 'on_leave': 0, 'on_duty': 0, 'not_marked': 0
        }
    else:
        staff_rows, staff_summary = _gather_attendance_data('Staff', report_date)

    return jsonify({
        "success": True,
        "date": report_date.strftime('%Y-%m-%d'),
        "students": {
            "summary": student_summary,
            "records": student_rows
        },
        "staff": {
            "summary": staff_summary,
            "records": staff_rows
        }
    })


@attendance_report_bp.route('/download/excel', methods=['GET'])
@jwt_required()
@role_required(['Admin', 'Staff'])
def download_excel():
    """
    Generates and returns an Excel file with student and/or staff attendance for today.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    report_date_str = request.args.get('date')
    role_filter = request.args.get('role') # 'Student' or 'Staff'
    
    current_user_id = get_jwt_identity()
    current_user = db.session.get(User, int(current_user_id))
    
    department_id = None
    if current_user.role == 'Staff':
        department_id = current_user.department_id
        # Staff can only download Student reports
        role_filter = 'Student'
        
    if report_date_str:
        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    else:
        report_date = date.today()

    wb = Workbook()

    # --- Style definitions ---
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    title_font = Font(bold=True, size=14, color="1A1A2E")
    summary_font = Font(bold=True, size=11, color="333333")
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    status_fills = {
        'Present': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'Absent': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'Late': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'On Leave': PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        'On Duty': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'Not Marked': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }

    def _write_section(ws, title, rows, summary, start_row=1):
        """Writes a titled attendance section into the worksheet."""
        row = start_row

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=f"{title} Attendance Report — {report_date.strftime('%d %B %Y')}")
        cell.font = title_font
        cell.alignment = left_align
        row += 1

        # College name
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value="Prathyusha Engineering College (PEC)")
        cell.font = Font(size=11, italic=True, color="666666")
        row += 2

        # Summary row
        summary_labels = [
            f"Total: {summary['total']}",
            f"Present: {summary['present']}",
            f"Absent: {summary['absent']}",
            f"Late: {summary['late']}",
            f"On Leave: {summary['on_leave']}",
            f"On Duty: {summary['on_duty']}",
            f"Not Marked: {summary['not_marked']}",
        ]
        for col_idx, label in enumerate(summary_labels, 1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = summary_font
            cell.alignment = center_align
            cell.border = thin_border
        row += 2

        # Table headers
        headers = ["S.No", "ID", "Username", "Department", "Status", "Check-In", "Check-Out"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1

        # Data rows
        for idx, r in enumerate(rows, 1):
            vals = [idx, r['id'], r['username'], r['department'], r['status'], r['check_in'], r['check_out']]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if col_idx == 5:  # Status column
                    cell.fill = status_fills.get(val, PatternFill())
            row += 1

        return row + 1  # return next available row

    def _setup_sheet_dimensions(ws):
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

    if role_filter == 'Student':
        ws = wb.active
        ws.title = "Student Attendance"
        _setup_sheet_dimensions(ws)
        student_rows, student_summary = _gather_attendance_data('Student', report_date, department_id)
        _write_section(ws, "Student", student_rows, student_summary)
        filename = f"PEC_Student_Attendance_{report_date.strftime('%Y-%m-%d')}.xlsx"
    elif role_filter == 'Staff':
        ws = wb.active
        ws.title = "Staff Attendance"
        _setup_sheet_dimensions(ws)
        staff_rows, staff_summary = _gather_attendance_data('Staff', report_date, department_id)
        _write_section(ws, "Staff", staff_rows, staff_summary)
        filename = f"PEC_Staff_Attendance_{report_date.strftime('%Y-%m-%d')}.xlsx"
    else:
        student_rows, student_summary = _gather_attendance_data('Student', report_date, department_id)
        staff_rows, staff_summary = _gather_attendance_data('Staff', report_date, department_id)
        # Both sections in separate sheets
        ws_students = wb.active
        ws_students.title = "Student Attendance"
        _setup_sheet_dimensions(ws_students)
        _write_section(ws_students, "Student", student_rows, student_summary)

        ws_staff = wb.create_sheet("Staff Attendance")
        _setup_sheet_dimensions(ws_staff)
        _write_section(ws_staff, "Staff", staff_rows, staff_summary)
        filename = f"PEC_Attendance_{report_date.strftime('%Y-%m-%d')}.xlsx"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@attendance_report_bp.route('/download/pdf', methods=['GET'])
@jwt_required()
@role_required(['Admin', 'Staff'])
def download_pdf():
    """
    Generates and returns a PDF file with student and/or staff attendance for today.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    report_date_str = request.args.get('date')
    role_filter = request.args.get('role') # 'Student' or 'Staff'
    
    current_user_id = get_jwt_identity()
    current_user = db.session.get(User, int(current_user_id))
    
    department_id = None
    if current_user.role == 'Staff':
        department_id = current_user.department_id
        # Staff can only download Student reports
        role_filter = 'Student'
        
    if report_date_str:
        try:
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
    else:
        report_date = date.today()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1A1A2E'),
        spaceAfter=4
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#666666'),
        spaceAfter=10,
        fontName='Helvetica-Oblique'
    )
    summary_style = ParagraphStyle(
        'SummaryText',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )

    elements = []

    def _add_section(title, rows, summary):
        elements.append(Paragraph(f"{title} Attendance Report — {report_date.strftime('%d %B %Y')}", title_style))
        elements.append(Paragraph("Prathyusha Engineering College (PEC)", subtitle_style))

        summary_text = (
            f"Total: {summary['total']} | "
            f"Present: {summary['present']} | "
            f"Absent: {summary['absent']} | "
            f"Late: {summary['late']} | "
            f"On Leave: {summary['on_leave']} | "
            f"On Duty: {summary['on_duty']} | "
            f"Not Marked: {summary['not_marked']}"
        )
        elements.append(Paragraph(summary_text, summary_style))

        # Build table data
        table_data = [['S.No', 'ID', 'Username', 'Department', 'Status', 'Check-In', 'Check-Out']]
        for idx, r in enumerate(rows, 1):
            table_data.append([
                str(idx), str(r['id']), r['username'], r['department'],
                r['status'], r['check_in'], r['check_out']
            ])

        if len(table_data) == 1:
            table_data.append(['', '', 'No records found', '', '', '', ''])

        col_widths = [40, 40, 140, 120, 90, 90, 90]
        table = Table(table_data, colWidths=col_widths)

        # Color-code statuses
        status_colors = {
            'Present': colors.HexColor('#C6EFCE'),
            'Absent': colors.HexColor('#FFC7CE'),
            'Late': colors.HexColor('#FFEB9C'),
            'On Leave': colors.HexColor('#D9E2F3'),
            'On Duty': colors.HexColor('#E2EFDA'),
            'Not Marked': colors.HexColor('#F2F2F2'),
        }

        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A1A2E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]

        # Apply status-specific background colors
        for row_idx in range(1, len(table_data)):
            status = table_data[row_idx][4]
            if status in status_colors:
                style_commands.append(
                    ('BACKGROUND', (4, row_idx), (4, row_idx), status_colors[status])
                )

        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        elements.append(Spacer(1, 20))

    if role_filter == 'Student':
        student_rows, student_summary = _gather_attendance_data('Student', report_date, department_id)
        _add_section("Student", student_rows, student_summary)
        filename = f"PEC_Student_Attendance_{report_date.strftime('%Y-%m-%d')}.pdf"
    elif role_filter == 'Staff':
        staff_rows, staff_summary = _gather_attendance_data('Staff', report_date, department_id)
        _add_section("Staff", staff_rows, staff_summary)
        filename = f"PEC_Staff_Attendance_{report_date.strftime('%Y-%m-%d')}.pdf"
    else:
        student_rows, student_summary = _gather_attendance_data('Student', report_date, department_id)
        staff_rows, staff_summary = _gather_attendance_data('Staff', report_date, department_id)
        
        _add_section("Student", student_rows, student_summary)
        from reportlab.platypus import PageBreak
        elements.append(PageBreak())
        _add_section("Staff", staff_rows, staff_summary)
        filename = f"PEC_Attendance_{report_date.strftime('%Y-%m-%d')}.pdf"

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )
